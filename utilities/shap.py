import shap
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd

# =========================================
# 1️⃣ SHAP values
# =========================================
def compute_shap_values(model, X):
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X)
    return explainer, shap_values


# =========================================
# 2️⃣ Colori per le categorical
# =========================================
def get_category_colors(feature_series):
    categories = feature_series.astype(str).unique()
    cmap = plt.get_cmap("tab10")
    color_dict = {cat: cmap(i % 10) for i, cat in enumerate(categories)}
    colors = feature_series.astype(str).map(color_dict)
    return colors, color_dict


# =========================================
# 3️⃣ Costruisci la LISTA colori (non matrice!)
# =========================================
def build_color_vector(X, cat_idx):
    """
    Restituisce un dizionario:
    { feature_name : vector_colors }
    Ogni vettore ha lunghezza n_samples.
    """
    color_vectors = {}
    categorical_cols = X.columns[cat_idx]

    for col in X.columns:
        if col in categorical_cols:
            colors, _ = get_category_colors(X[col])
            color_vectors[col] = colors.values
        else:
            vals = X[col].values
            mn, mx = np.nanmin(vals), np.nanmax(vals)
            nv = (vals - mn) / (mx - mn + 1e-8)
            color_vectors[col] = np.array([plt.cm.RdBu_r(v) for v in nv])

    return color_vectors


# =========================================
# 4️⃣ Summary plot UNICO con categorical colorate
# =========================================
def save_shap_summary_custom(shap_values, X, color_vectors, output_path):
    """
    Passiamo a SHAP una funzione di lookup dei colori per ogni feature.
    """
    def color_lookup(values, feature_name):
        return color_vectors[feature_name]

    plt.figure()
    shap.summary_plot(
        shap_values,
        X,
        feature_names=X.columns,
        show=False,
        color=color_lookup   # ← MAGIA!
    )
    plt.tight_layout()
    plt.savefig(output_path)
    plt.close()
    print(f"Salvato summary plot con categorical colorate → {output_path}")


# =========================================
# 5️⃣ Plot separati per ogni categorical
# =========================================
def plot_shap_categorical_manual(shap_values, X, cat_idx, folder):
    categorical_cols = X.columns[cat_idx]
    colors_dict = {}

    for col in categorical_cols:
        values = X[col].astype(str)
        categories = values.unique()
        cmap = plt.get_cmap("tab10")

        color_map = {cat: cmap(i % 10) for i, cat in enumerate(categories)}
        colors_dict[col] = color_map

        col_idx = X.columns.get_loc(col)
        shap_col = shap_values[:, col_idx]

        plt.figure(figsize=(6,4))
        for cat in categories:
            idx = values == cat
            plt.scatter(shap_col[idx],
                        np.random.uniform(0,1,size=sum(idx)),
                        c=[color_map[cat]],
                        alpha=0.7,
                        label=cat)

        plt.title(f"SHAP categorical: {col}")

        col = col.replace('/', '')
        plt.yticks([])
        plt.legend(bbox_to_anchor=(1.05,1), loc="upper left")
        plt.tight_layout()
        plt.savefig(os.path.join(folder, f"shap_categorical_{col}.png"))
        plt.close()

    return colors_dict


# =========================================
# 6️⃣ Legenda Excel
# =========================================
def save_categorical_legend_excel(colors_dict, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorical Legend"

    row = 1
    for col, cmap in colors_dict.items():
        ws.cell(row=row, column=1, value=col)
        row += 1

        for cat, color in cmap.items():
            cell = ws.cell(row=row, column=1, value=str(cat))

            # converti da (r,g,b,a?) → hex
            r, g, b = [int(255 * x) for x in color[:3]]
            hex_color = f"{r:02X}{g:02X}{b:02X}"

            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            row += 1

        row += 1  # spazio tra le feature

    wb.save(output_path)
    print(f"Legenda Excel salvata → {output_path}")



def build_categorical_color_dict(X, cat_features_idx, cmap_name="tab10"):
    """
    Genera un dizionario completo:
    { colname : { category: color } }
    """
    categorical_cols = X.columns[cat_features_idx]
    cmap = plt.get_cmap(cmap_name)

    colors_dict = {}

    for col in categorical_cols:
        values = X[col].astype("object")
        unique_vals = values.dropna().unique()

        color_map = {}
        for i, cat in enumerate(unique_vals):
            color_map[cat] = cmap(i % 10)

        # colore per missing
        color_map["missing"] = (0.5, 0.5, 0.5)

        colors_dict[col] = color_map

    return colors_dict



# =========================================
# 7️⃣ Funzione unica
# =========================================
def generate_shap_plots(model, X, cat_features_idx, folder):
    os.makedirs(folder, exist_ok=True)

    # SHAP values
    _, shap_values = compute_shap_values(model, X)
    filename = os.path.join(folder, "shap_summary_all.png")

    # Colori uniformi
    color_vectors = build_color_vector(X, cat_features_idx)
    # Summary SHAP unico
    save_shap_summary_custom(
        shap_values,
        X,
        color_vectors,
        filename
    )

    # Plot extra per categorical
    colors_dict = plot_shap_categorical_manual(shap_values, X, cat_features_idx, folder)

    # Legenda Excel
    save_categorical_legend_excel(colors_dict, os.path.join(folder, "shap_legend.xlsx"))

def shap_summary_custom(shap_values, X, colors_dict, output_path=None):
    feature_names = X.columns
    n_features = len(feature_names)

    plt.figure(figsize=(10, 0.35*n_features + 2))

    # Ordina per importanza media SHAP
    mean_abs_shap = np.abs(shap_values).mean(axis=0)
    order = np.argsort(mean_abs_shap)[::-1]

    for pos, idx in enumerate(order):
        f = feature_names[idx]
        shap_f = shap_values[:, idx]
        x_f = X[f]

        y = np.ones(len(shap_f))*pos

        # 🔵 Se è numerica → gradiente (tipo SHAP)
        if np.issubdtype(x_f.dtype, np.number):
            colors = plt.cm.coolwarm(
                (x_f - x_f.min()) / (x_f.max() - x_f.min() + 1e-9)
            )

        # 🟢 Se è categoriale → usa colori_dict
        else:
            colors = x_f.map(colors_dict.get(f, {}))
            colors = colors.fillna("#808080")  # eventuali non mappati

        plt.scatter(shap_f, y, color=colors, s=18, alpha=0.8, edgecolor='none')

    plt.yticks(range(n_features), feature_names[order])
    plt.xlabel("SHAP value")
    plt.title("Custom SHAP Summary Plot (colori categoriali inclusi)")
    plt.tight_layout()

    if output_path:
        plt.savefig(output_path, dpi=200)
        print(f"Plot salvato in {output_path}")

    plt.show()


def save_shap_values_csv(model, X, output_path="shap_values.csv"):
    """
    Calcola i valori SHAP per ogni campione e li salva in un CSV.

    Parameters
    ----------
    model : fitted model
        Il modello già allenato, compatibile con shap.TreeExplainer
    X : pd.DataFrame o np.array
        Il dataset usato per calcolare i valori shap
    cat_features_idx : list o None
        Indici delle variabili categoriche (serve solo per CatBoost)
    output_path : str
        Percorso del CSV di output
    """

    # SHAP explainer
    try:
        explainer = shap.TreeExplainer(model)
    except Exception:
        explainer = shap.KernelExplainer(model.predict, X.sample(100))  # fallback molto più lento

    # Compute SHAP values
    shap_values = explainer.shap_values(X)

    # Se il modello è binario, shap_values è una lista (prendo la classe 1)
    if isinstance(shap_values, list):
        shap_values = shap_values[1]

    # Converto in DataFrame
    shap_df = pd.DataFrame(shap_values, columns=X.columns)

    # Salvo
    shap_df.to_csv(output_path, index=False)

    print(f"SHAP values salvati in: {output_path}")
    return shap_df